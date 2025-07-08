import os
import threading
import re
import requests
import pandas as pd
import csv
from datetime import datetime, timezone
from io import BytesIO

from flask import Flask
from telegram import Update, ReplyKeyboardMarkup, KeyboardButton
from telegram.ext import (
    ApplicationBuilder,
    CommandHandler,
    MessageHandler,
    filters,
    ContextTypes,
)
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from config import (
    TOKEN, SELF_URL, PORT,
    BRANCH_URLS, NOTIFY_URLS,
    NOTIFY_LOG_FILE_UG, NOTIFY_LOG_FILE_RK
)
from zones import normalize_sheet_url, load_zones

# Сопоставление «сырого» названия филиала из CSV в ключи BRANCH_URLS
BRANCH_KEY_MAP = {
    "Тимашевский":      "Тимашевские ЭС",
    "Усть-Лабинский":   "Усть-Лабинские ЭС",
    "Тихорецкий":       "Тихорецкие ЭС",
    "Сочинский":        "Сочинские ЭС",
    "Славянский":       "Славянские ЭС",
    "Ленинградский":    "Ленинградские ЭС",
    "Лабинский":        "Лабинские ЭС",
    "Краснодарский":    "Краснодарские ЭС",
    "Армавирский":      "Армавирские ЭС",
    "Адыгейский":       "Адыгейские ЭС",
    "Центральный":      "Центральные ЭС",
    "Западный":         "Западные ЭС",
    "Восточный":        "Восточные ЭС",
    "Южный":            "Южные ЭС",
    "Северо-Восточный": "Северо-Восточные ЭС",
    "Юго-Восточный":    "Юго-Восточные ЭС",
    "Северный":         "Северные ЭС",
}

app = Flask(__name__)
application = ApplicationBuilder().token(TOKEN).build()

# Инициализация CSV-файлов для логов уведомлений
for lf in (NOTIFY_LOG_FILE_UG, NOTIFY_LOG_FILE_RK):
    if not os.path.exists(lf):
        with open(lf, "w", newline="", encoding="utf-8") as f:
            csv.writer(f).writerow([
                "Filial","РЭС",
                "SenderID","SenderName",
                "RecipientID","RecipientName",
                "Timestamp","Coordinates"
            ])

# Клавиатуры
kb_back = ReplyKeyboardMarkup([["🔙 Назад"]], resize_keyboard=True)
kb_actions = ReplyKeyboardMarkup(
    [["🔍 Поиск по ТП"], ["🔔 Отправить уведомление"], ["🔙 Назад"]],
    resize_keyboard=True
)
kb_request_location = ReplyKeyboardMarkup(
    [[KeyboardButton("📍 Отправить геолокацию", request_location=True)], ["🔙 Назад"]],
    resize_keyboard=True
)

def build_initial_kb(vis_flag: str, res_flag: str) -> ReplyKeyboardMarkup:
    f = vis_flag.strip().upper()
    if f == "ALL":
        nets = ["⚡ Россети ЮГ", "⚡ Россети Кубань"]
    elif f == "UG":
        nets = ["⚡ Россети ЮГ"]
    else:
        nets = ["⚡ Россети Кубань"]
    buttons = [[n] for n in nets]
    buttons.append(["📞 Телефоны провайдеров"])
    if res_flag.strip().upper() == "ALL":
        buttons.append(["📝 Сформировать отчёт"])
    return ReplyKeyboardMarkup(buttons, resize_keyboard=True)

def build_report_kb(vis_flag: str) -> ReplyKeyboardMarkup:
    f = vis_flag.strip().upper()
    rows = []
    if f in ("ALL", "UG"):
        rows.append(["📊 Уведомления о бездоговорных ВОЛС ЮГ"])
    if f in ("ALL", "RK"):
        rows.append(["📊 Уведомления о бездоговорных ВОЛС Кубань"])
    rows += [["📋 Выгрузить информацию по контрагентам"], ["🔙 Назад"]]
    return ReplyKeyboardMarkup(rows, resize_keyboard=True)

# === /start ===
async def start_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    vis_map, raw_branch_map, res_map, names, resp_map = load_zones()
    if uid not in raw_branch_map:
        await update.message.reply_text("🚫 У вас нет доступа.", reply_markup=kb_back)
        return

    raw = raw_branch_map[uid]
    branch_key = "All" if raw == "All" else BRANCH_KEY_MAP.get(raw, raw)

    context.user_data.clear()
    context.user_data.update({
        "step":        "INIT",
        "vis_flag":    vis_map[uid],
        "branch_user": branch_key,
        "res_user":    res_map[uid],
        "name":        names[uid],
        "resp_map":    resp_map
    })

    await update.message.reply_text(
        f"👋 Приветствую Вас, {names[uid]}! Выберите опцию:",
        reply_markup=build_initial_kb(vis_map[uid], res_map[uid])
    )

# === TEXT handler ===
async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    if "step" not in context.user_data:
        return await start_cmd(update, context)

    vis_map, raw_branch_map, res_map, names, resp_map = load_zones()
    step        = context.user_data["step"]
    vis_flag    = context.user_data["vis_flag"]
    branch_user = context.user_data["branch_user"]
    res_user    = context.user_data["res_user"]
    name        = context.user_data["name"]

    # «Назад»
    if text == "🔙 Назад":
        if step in ("AWAIT_TP_INPUT","DISAMB","NOTIFY_AWAIT_TP","NOTIFY_DISAMB","NOTIFY_VL"):
            context.user_data["step"] = "BRANCH"
            await update.message.reply_text("Выберите действие:", reply_markup=kb_actions)
            return
        if step == "REPORT_MENU":
            context.user_data["step"] = "INIT"
            await update.message.reply_text(
                f"👋 Приветствую Вас, {name}! Выберите опцию:",
                reply_markup=build_initial_kb(vis_flag, res_user)
            )
            return
        context.user_data["step"] = "INIT"
        await update.message.reply_text(
            "Выберите опцию:",
            reply_markup=build_initial_kb(vis_flag, res_user)
        )
        return

    # INIT
    if step == "INIT":
        if text == "📞 Телефоны провайдеров":
            context.user_data["step"] = "VIEW_PHONES"
            await update.message.reply_text("📞 Телефоны провайдеров:\n…", reply_markup=kb_back)
            return

        if text == "📝 Сформировать отчёт":
            if res_user.strip().upper() != "ALL":
                await update.message.reply_text(
                    f"{name}, выгрузка отчётов доступна только при общем доступе по РЭС.",
                    reply_markup=build_initial_kb(vis_flag, res_user)
                )
                return
            context.user_data["step"] = "REPORT_MENU"
            await update.message.reply_text(
                "📝 Выберите тип отчёта:",
                reply_markup=build_report_kb(vis_flag)
            )
            return

        allowed = (
            ["⚡ Россети ЮГ","⚡ Россети Кубань"] if vis_flag=="All"
            else ["⚡ Россети ЮГ"] if vis_flag=="UG"
            else ["⚡ Россети Кубань"]
        )
        if text not in allowed:
            await update.message.reply_text(
                f"{name}, доступны: {', '.join(allowed)}",
                reply_markup=build_initial_kb(vis_flag, res_user)
            )
            return

        selected_net = text.replace("⚡ ","")
        context.user_data.update({"step":"NET","net":selected_net})
        if branch_user != "All":
            branches = [branch_user]
        else:
            branches = list(BRANCH_URLS[selected_net].keys())
        kb = ReplyKeyboardMarkup([[b] for b in branches] + [["🔙 Назад"]], resize_keyboard=True)
        await update.message.reply_text("Выберите филиал:", reply_markup=kb)
        return

    # REPORT_MENU
    if step == "REPORT_MENU":
        if text in ("📊 Уведомления о бездоговорных ВОЛС ЮГ", "📊 Уведомления о бездоговорных ВОЛС Кубань"):
            log_file = NOTIFY_LOG_FILE_UG if text.endswith("ЮГ") else NOTIFY_LOG_FILE_RK
            df = pd.read_csv(log_file)
            bio = BytesIO()
            with pd.ExcelWriter(bio, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, sheet_name="Логи")
                ws = writer.sheets["Логи"]
                pink = PatternFill(fill_type="solid", start_color="FFF4F4", end_color="FFF4F4")
                for col_idx in range(1, len(df.columns)+1):
                    ws.cell(row=1, column=col_idx).fill = pink
                for idx, col in enumerate(df.columns, start=1):
                    max_len = max(df[col].astype(str).map(len).max(), len(col))
                    ws.column_dimensions[get_column_letter(idx)].width = max_len + 2
            bio.seek(0)
            fname = "log_ug.xlsx" if text.endswith("ЮГ") else "log_rk.xlsx"
            await update.message.reply_document(bio, filename=fname)
        elif text == "📋 Выгрузить информацию по контрагентам":
            await update.message.reply_text(
                "📋 Справочник контрагентов — скоро будет!",
                reply_markup=build_report_kb(vis_flag)
            )
        else:
            await update.message.reply_text(
                "📝 Выберите тип отчёта:", reply_markup=build_report_kb(vis_flag)
            )
        return

    # NET → филиал
    if step == "NET":
        selected_net = context.user_data["net"]
        if branch_user!="All" and text!=branch_user:
            await update.message.reply_text(
                f"{name}, доступен только филиал «{branch_user}».", reply_markup=kb_back
            )
            return
        if text not in BRANCH_URLS[selected_net]:
            await update.message.reply_text(
                f"⚠ Филиал «{text}» не найден.", reply_markup=kb_back
            )
            return
        context.user_data.update({"step":"BRANCH","branch":text})
        await update.message.reply_text("Выберите действие:", reply_markup=kb_actions)
        return

    # BRANCH → действие
    if step == "BRANCH":
        if text == "🔍 Поиск по ТП":
            context.user_data["step"] = "AWAIT_TP_INPUT"
            await update.message.reply_text("Введите номер ТП:", reply_markup=kb_back)
            return
        if text == "🔔 Отправить уведомление":
            context.user_data["step"] = "NOTIFY_AWAIT_TP"
            await update.message.reply_text("Введите номер ТП для уведомления:", reply_markup=kb_back)
            return

    # AWAIT_TP_INPUT
    if step == "AWAIT_TP_INPUT":
        net    = context.user_data["net"]
        branch = context.user_data["branch"]
        url    = BRANCH_URLS[net].get(branch, "").strip()
        if not url:
            await update.message.reply_text(
                f"⚠️ URL для «{branch}» не настроен.", reply_markup=kb_back
            )
            context.user_data["step"] = "BRANCH"
            return
        try:
            df = pd.read_csv(normalize_sheet_url(url))
        except Exception as e:
            await update.message.reply_text(f"❌ Ошибка загрузки: {e}", reply_markup=kb_back)
            context.user_data["step"] = "BRANCH"
            return

        if res_user.upper() != "ALL":
            df = df[df["РЭС"].str.upper() == res_user.upper()]

        df["D_UP"] = df["Наименование ТП"].str.upper().str.replace(r"\W","",regex=True)
        q = re.sub(r"\W","", text.upper())
        found = df[df["D_UP"].str.contains(q, na=False)]

        if found.empty:
            await update.message.reply_text(
                f"В {res_user} отсутствуют ТП, удовлетворяющие параметры поиска.",
                reply_markup=kb_back
            )
            context.user_data["step"] = "BRANCH"
            return

        ulist = found["Наименование ТП"].unique().tolist()
        if len(ulist) > 1:
            context.user_data.update({"step":"DISAMB","amb_df":found})
            kb = ReplyKeyboardMarkup([[tp] for tp in ulist] + [["🔙 Назад"]], resize_keyboard=True)
            await update.message.reply_text("Выберите ТП:", reply_markup=kb)
            return

        tp = ulist[0]
        det=found[found["Наименование ТП"]==tp]
        resname=det.iloc[0]["РЭС"]
        await update.message.reply_text(
            f"{resname}, {tp} ({len(det)}) ВОЛС с договором аренды:", reply_markup=kb_actions
        )
        for _,r in det.iterrows():
            await update.message.reply_text(
                f"📍 ВЛ {r['Уровень напряжения']} {r['Наименование ВЛ']}\n"
                f"Опоры: {r['Опоры']}\n"
                f"Провайдер: {r.get('Наименование Провайдера','')}",
                reply_markup=kb_actions
            )
        context.user_data["step"] = "BRANCH"
        return

    # DISAMB
    if step == "DISAMB":
        if text == "🔙 Назад":
            context.user_data["step"] = "AWAIT_TP_INPUT"
            await update.message.reply_text("Введите номер ТП:", reply_markup=kb_back)
            return
        found = context.user_data["amb_df"]
        if text not in found["Наименование ТП"].unique():
            return
        det = found[found["Наименование ТП"]==text]
        resname=det.iloc[0]["РЭС"]
        await update.message.reply_text(
            f"{resname}, {text} ({len(det)}) ВОЛС с договором аренды:", reply_markup=kb_actions
        )
        for _,r in det.iterrows():
            await update.message.reply_text(
                f"📍 ВЛ {r['Уровень напряжения']} {r['Наименование ВЛ']}\n"
                f"Опоры: {r['Опоры']}\n"
                f"Провайдер: {r.get('Наименование Провайдера','')}",
                reply_markup=kb_actions
            )
        context.user_data["step"] = "BRANCH"
        return

    # Блоки уведомлений (NOTIFY_AWAIT_TP, NOTIFY_DISAMB, NOTIFY_VL, location_handler) — без изменений

# Обработчик геолокации
async def location_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if context.user_data.get("step") != "NOTIFY_GEO":
        return
    loc       = update.message.location
    tp        = context.user_data["tp"]
    vl        = context.user_data["vl"]
    res_tp    = context.user_data["notify_res"]
    sender    = context.user_data["name"]
    _,_,_,names,resp_map = load_zones()

    recipients = [
        uid for uid, r in resp_map.items()
        if r and r.strip().lower() == res_tp.strip().lower()
    ]

    msg   = f"🔔 Уведомление от {sender}, {res_tp} РЭС, {tp}, {vl} – Найден бездоговорной ВОЛС"
    log_f = NOTIFY_LOG_FILE_UG if context.user_data["net"]=="Россети ЮГ" else NOTIFY_LOG_FILE_RK

    for cid in recipients:
        await context.bot.send_message(cid, msg)
        await context.bot.send_location(cid, loc.latitude, loc.longitude)
        await context.bot.send_message(cid, f"📍 Широта: {loc.latitude:.6f}, Долгота: {loc.longitude:.6f}")
        with open(log_f, "a", newline="", encoding="utf-8") as f:
            csv.writer(f).writerow([
                context.user_data["branch"],
                res_tp,
                update.effective_user.id,
                sender,
                cid,
                names.get(cid, ""),
                datetime.now(timezone.utc).isoformat(),
                f"{loc.latitude:.6f},{loc.longitude:.6f}"
            ])

    if recipients:
        names_list = [names[c] for c in recipients]
        await update.message.reply_text(
            f"✅ Уведомление отправлено: {', '.join(names_list)}",
            reply_markup=kb_actions
        )
    else:
        await update.message.reply_text(
            f"⚠ Ответственный на {res_tp} РЭС не назначен.",
            reply_markup=kb_actions
        )

    context.user_data["step"] = "BRANCH"

# Регистрируем хендлеры
application.add_handler(CommandHandler("start", start_cmd))
application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text))
application.add_handler(MessageHandler(filters.LOCATION, location_handler))

if __name__ == "__main__":
    if SELF_URL:
        threading.Thread(target=lambda: requests.get(f"{SELF_URL}/webhook"), daemon=True).start()
    application.run_webhook(
        listen="0.0.0.0", port=PORT,
        url_path="webhook", webhook_url=f"{SELF_URL}/webhook"
    )
